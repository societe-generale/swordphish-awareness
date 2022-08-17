import datetime
from django.utils.timezone import get_current_timezone
import uuid
from base64 import b64decode
from smtplib import SMTPRecipientsRefused, SMTPServerDisconnected
from email.mime.image import MIMEImage
from email.header import Header
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Alignment
from ckeditor.fields import RichTextField
from django.db import models, transaction
from django.core.validators import validate_email
from django.core.exceptions import ObjectDoesNotExist
from django.forms import ValidationError
from django.core import mail
from django.conf import settings
from django.urls import reverse
from django.utils.translation import gettext as _
from django.core.validators import RegexValidator

from LocalUsers.models import SwordphishUser, get_admin
from Main.utils import send_alert_new_campaign

import re
import logging
from time import sleep
# Create your models here.

class Attribute(models.Model):
    key = models.CharField(db_index=True, max_length=240)
    value = models.CharField(db_index=True, max_length=240)

    def __str__(self):
        return "%s => %s" % (self.key, self.value)

    class Meta:
        unique_together = ('key', 'value')


class Target(models.Model):
    mail_address = models.EmailField(db_index=True,)
    attributes = models.ManyToManyField(Attribute)

    def __str__(self):
        return self.mail_address

    def addAttribute(self, attribute, value):
        l = self.attributes.filter(key=attribute, value=value)
        if not l:
            newAtt = None
            try:
                newAtt = Attribute.objects.get(key=attribute, value=value)
            except ObjectDoesNotExist:
                newAtt = Attribute.objects.create(key=attribute, value=value)
            self.attributes.add(newAtt)
            self.save()

    def removeAttribute(self, attribute):
        self.attributes.remove(attribute)
        self.save()



class TargetList(models.Model):
    class Meta:
        ordering = ["name", "-creation_date"]

    name = models.CharField(max_length=240)
    targets = models.ManyToManyField(Target)
    author = models.ForeignKey(SwordphishUser, null=True, on_delete=models.SET(get_admin))
    creation_date = models.DateTimeField(blank=True, default=datetime.datetime.now)

    def __str__(self):
        return self.name

    def loadTargetsFromUploadedXslx(self, uploadedFile):
        wb = load_workbook(filename=uploadedFile, read_only=True)
        ws = wb.active
        targets = {}
        header = next(ws.rows)
        tags_keys = []
        n = 1
        for x in header[1:]:
            if x.value is not None:
                tags_keys.append('ORDN-' +  "%03d" % n + '-' + x.value)
                n += 1
        for row in ws.iter_rows(row_offset=1):
            if row[0] and row[0].value:
                email = row[0].value.lower()
                tags_values = [x.value if x.value is not None else "N/A" for x in row[1:]]
                tags = dict(zip(tags_keys, tags_values))
                try:
                    validate_email(email)
                except ValidationError:
                    raise ValidationError(email)
                targets[email] = tags
        new_targets = []
        with transaction.atomic():
            for email in targets:
                if not self.targets.filter(mail_address=email).exists():
                    tmp = Target.objects.create(mail_address=email)
                    atts = []
                    for tag, value in targets[email].items():
                        try:
                            newatt = Attribute.objects.get(key=tag, value=value)
                        except ObjectDoesNotExist:
                            newatt = Attribute.objects.create(key=tag, value=value)
                        atts.append(newatt)
                    tmp.attributes.add(*atts)
                    new_targets.append(tmp)
            self.targets.add(*new_targets)

        transaction.commit()

        return True

    def removeTarget(self, target):
        self.targets.remove(target)
        self.targets.filter(mail_address=target.mail_address).delete()

    def addTarget(self, email, attributes):
        newTarget = Target(mail_address=email)
        newTarget.save()
        for att, value in attributes.iteritems():
            newTarget.addAttribute(att, value)

    def is_used(self):
        return len(Campaign.objects.filter(targets=self.id)) > 0

    def export_to_xlsx(self):
        targets = self.targets.all()
        wb = Workbook()
        ws = wb.active
        tags = set()
        mails = {}
        for target in targets:
            mails[target.mail_address] = {}
            attributes = target.attributes.all()
            for attribute in attributes:
                mails[target.mail_address][attribute.key] = attribute.value
                tags.add(attribute.key)
        raw_header = list(sorted(tags))
        header = []
        for h in raw_header:
            h = re.sub(r'ORDN-[0-9]{3}-','',h)
            header.append(h)
        header.insert(0, "email")
        ws.append(header)
        for address in mails:
            line = []
            line.append(address)
            for tag in sorted(tags):
                if tag in mails[address]:
                    line.append(mails[address][tag])
                else:
                    line.append("N/A")
            ws.append(line)

        c = ws['B2']
        ws.freeze_panes = c
        head = ws[1]
        ft = Font(bold=True)
        al = Alignment(horizontal="center", vertical="center")
        for cell in head:
            cell.font = ft
        for row in ws.rows:
            for cell in row:
                cell.alignment = al
        return save_virtual_workbook(wb)



class AnonymousTarget(models.Model):
    uniqueid = models.CharField(db_index=True, max_length=36, default=uuid.uuid4)
    attributes = models.ManyToManyField(Attribute)
    mail_sent_time = models.DateTimeField(blank=True, null=True)
    mail_opened = models.BooleanField(default=False)
    mail_opened_time = models.DateTimeField(blank=True, null=True)
    link_clicked = models.BooleanField(default=False)
    link_clicked_time = models.DateTimeField(blank=True, null=True)
    attachment_opened = models.BooleanField(default=False)
    attachment_opened_time = models.DateTimeField(blank=True, null=True)
    form_submitted = models.BooleanField(default=False)
    form_submitted_time = models.DateTimeField(blank=True, null=True)
    reported = models.BooleanField(default=False)
    reported_time = models.DateTimeField(blank=True, null=True)

    def __str__(self):
        return self.uniqueid

    def addAttribute(self, attribute, value):
        l = self.attributes.filter(key=attribute, value=value)
        self.attributes.add(l[0])

    def importAttributes(self, target):
        l = target.attributes.all()
        for att in l:
            self.attributes.add(att)



class PhishmailDomain(models.Model):
    class Meta:
        ordering = ["domain"]

    domain = models.CharField(max_length=50)

    def __str__(self):
        return self.domain



class Template(models.Model):

    class Meta:
        ordering = ["name", "-creation_date"]

    TEMPLATE_TYPE = [
        ('#1', _('Mail with link')),
        ('1', _('Mail with link template')),
        ('#2', _('Mail with attachment')),
        ('2', _('Mail with attachment template')),
        ('3', _('Attachment template')),
        ('#3', _('Action after click')),
        ('4', _('Redirection')),
        ('5', _('Awareness template')),
        ('6', _('Fake form template')),
        ('7', _('Fake ransomware template')),
    ]

    template_type = models.CharField(max_length=2, choices=TEMPLATE_TYPE, default=1)
    name = models.CharField(max_length=200)
    author = models.ForeignKey(SwordphishUser, on_delete=models.SET(get_admin))
    creation_date = models.DateTimeField(auto_now_add=True, blank=True)
    public = models.BooleanField(default=False)
    title = models.CharField(max_length=100)
    timeout = models.PositiveIntegerField(default=10)
    text = RichTextField()

    def __str__(self):
        return self.name

    def is_used(self):
        if self.template_type == "1" or self.template_type == "2":
            return bool(Campaign.objects.filter(mail_template=self.id))
        if self.template_type == "3":
            return bool(Campaign.objects.filter(attachment_template=self.id))
        if self.template_type == "4" or self.template_type == "5":
            return bool(Campaign.objects.filter(onclick_action=self.id))
        if self.template_type == "6":
            return bool(Campaign.objects.filter(fake_form=self.id))
        if self.template_type == "7":
            return bool(Campaign.objects.filter(fake_ransom=self.id))

        return False



class Campaign(models.Model):

    class Meta:
        ordering = ["-start_date", "-end_date", "creation_date"]

    CAMPAIGN_TYPES = [
        ("1", _("Simple")),
        ("2", _("With Attachment")),
        ("3", _("Fake Form")),
        ("4", _("Fake Ransomware"))
    ]

    CAMPAIGN_STATUS = [
        ('1', _("Not Started")),
        ('2', _("Running")),
        ('3', _("Finished")),
    ]

    campaign_type = models.CharField(max_length=1, choices=CAMPAIGN_TYPES, default="1")

    name = models.CharField(max_length=200)

    author = models.ForeignKey(SwordphishUser, on_delete=models.SET(get_admin))

    creation_date = models.DateTimeField(auto_now_add=True, blank=True)

    start_date = models.DateTimeField()

    end_date = models.DateTimeField()

    targets = models.ManyToManyField(TargetList)

    targets_count = models.IntegerField(default=0)

    anonymous_targets = models.ManyToManyField(AnonymousTarget)

    status = models.CharField(max_length=1, choices=CAMPAIGN_STATUS, default="1")

    mail_template = models.ForeignKey(Template,
                                      related_name='%(class)s_mail',
                                      on_delete=models.PROTECT)

    from_name = models.CharField(max_length=50)
    from_domain = models.ForeignKey(PhishmailDomain,
                                    related_name='%(class)s_from_domain',
                                    on_delete=models.PROTECT)

    displayname_regex = RegexValidator(regex=r'^[^@]{1,100}$',
                                       message=_("Display name must not contains @"))

    display_name = models.CharField(validators=[displayname_regex],
                                    max_length=100,
                                    null=True,
                                    blank="True")

    attachment_template = models.ForeignKey(Template,
                                            default=None,
                                            blank=True,
                                            null=True,
                                            related_name='%(class)s_attachment',
                                            on_delete=models.PROTECT)
    fake_form = models.ForeignKey(Template,
                                  default=None,
                                  blank=True,
                                  null=True,
                                  related_name='%(class)s_fake_form',
                                  on_delete=models.PROTECT)

    fake_ransom = models.ForeignKey(Template,
                                    default=None,
                                    blank=True,
                                    null=True,
                                    related_name='%(class)s_fake_ransom',
                                    on_delete=models.PROTECT)

    onclick_action = models.ForeignKey(Template,
                                       default=None,
                                       blank=True,
                                       null=True,
                                       related_name='%(class)s_action',
                                       on_delete=models.PROTECT)

    host_subdomain_regex = RegexValidator(regex=r'^[-.a-z0-9]{0,200}(?<!\.)$',
                                          message=_("Only a-z, 0-9, - and ., must not end with a .")
                                          )

    host_subdomain = models.CharField(validators=[host_subdomain_regex],
                                      max_length=200,
                                      blank=True,
                                      null=True)

    host_domain = models.ForeignKey(PhishmailDomain,
                                    default=None,
                                    blank=True,
                                    null=True,
                                    related_name='%(class)s_host_domain',
                                    on_delete=models.PROTECT)

    enable_mail_tracker = models.BooleanField(default=True)

    enable_attachment_tracker = models.BooleanField(default=True)

    testid = models.CharField(max_length=36, default=uuid.uuid4)

    def __str__(self):
        return self.name

    def __buildemail(self):
        i = 0
        soup = BeautifulSoup(self.mail_template.text, "html.parser")
        imgs = []
        for img in soup.findAll("img"):
            if not img.has_attr("src") or not img["src"][:5].lower() == "data:":
                continue
            tmp = img["src"].split(',')
            if len(tmp) > 1:
                base64 = tmp[1]
                mime = tmp[0].split(';')[0].split(':')[1]
                mimeimg = MIMEImage(b64decode(base64))
                mimeimg.set_type(mime)
                mimeimg.add_header('Content-ID', '<img%s>' % i)
                imgs.append(mimeimg)
            img["src"] = "cid:img%s" % i
            if "style" in img:
                styles = img["style"].split(";")
                height = 0
                width = 0
                for style in styles:
                    if "height" in style:
                        tmp = style.split(":")[1].replace("px", "")
                        height = str(int(tmp) * 75 / 100)
                    if "width" in style:
                        tmp = style.split(":")[1].replace("px", "")
                        width = str(int(tmp) * 75 / 100)
                if height != 0:
                    img["height"] = height
                if width != 0:
                    img["width"] = width
            i += 1

        if self.enable_mail_tracker:
            if not soup.findAll("img", attrs={"src": "FIXMEMAILTRACKER"}):
                img = soup.new_tag("img", src="FIXMEMAILTRACKER")
                soup.body.append(img)

        return {"text": str(soup), "imgs": imgs}

    def __buildattachment(self):
        if self.campaign_type != "2":
            return ""
        i = 0
        soup = BeautifulSoup(self.attachment_template.text, "html.parser")
        imgs = []
        for img in soup.findAll("img"):
            if not img.has_attr("src") or not img["src"][:5].lower() == "data:":
                continue
            tmp = img["src"].split(',')
            mime = tmp[0].split(';')[0].split(':')[1]
            mimeimg = tmp[1]
            imgs.append((mimeimg, mime, "fichier_files/img%s" % i))
            img["src"] = "fichier_files/img%s" % i
            if "style" in img:
                styles = img["style"].split(";")
                height = 0
                width = 0
                for style in styles:
                    if "height" in style:
                        tmp = style.split(":")[1].replace("px", "")
                        height = str(int(tmp) * 75 / 100)
                    if "width" in style:
                        tmp = style.split(":")[1].replace("px", "")
                        width = str(int(tmp) * 75 / 100)
                if height != 0:
                    img["height"] = height
                if width != 0:
                    img["width"] = width
            i += 1

        if self.enable_attachment_tracker:
            if not soup.findAll("img", attrs={"src": "FIXMEDOCTRACKER"}):
                img = soup.new_tag("img", src="FIXMEDOCTRACKER")
                soup.body.append(img)

        if not soup.findAll("meta", attrs={"charset": "UTF-8"}):
            charset = soup.new_tag("meta", charset="UTF-8")
            soup.head.append(charset)

        result = ""
        result += "MIME-Version: 1.0\n"
        result += 'Content-Type: multipart/related; boundary="----=_SwordPhish_Decoy_000_001"\n\n'
        result += "------=_SwordPhish_Decoy_000_001\n"
        result += "Content-Location: file:///C:/fichier.html\n"
        result += 'Content-Type: text/html; charset="utf-8"\n\n'
        result += "%s\n\n\n" % soup
        for img in imgs:
            result += "------=_SwordPhish_Decoy_000_001\n"
            result += "Content-Location: file:///C:/%s\n" % img[2]
            result += "Content-Type: %s\n" % img[1]
            result += "Content-Transfer-Encoding: base64\n\n"
            result += img[0]
            result += "\n\n\n"
        result += "------=_SwordPhish_Decoy_000_001--\n"

        return result

    def __sendemail(self, recipient, targetid, connec, mail_content):
        base_mail = mail_content
        mail_content = "Read this mail with a HTML compatible client"
        sender = self.from_name + '@' + self.from_domain.domain
        try:
            target = AnonymousTarget.objects.get(uniqueid=targetid)
        except ObjectDoesNotExist:
            target = None
        if self.display_name != "":
            from_mail = "%s <%s>" % (self.display_name, sender)
        else:
            from_mail = sender

        email = mail.EmailMultiAlternatives(self.mail_template.title,
                                            mail_content,
                                            from_mail,
                                            [recipient],
                                            connection=connec)

        if self.host_subdomain is not None and self.host_subdomain != "":
            linkurl = "%s.%s%s" % (self.host_subdomain,
                                   self.host_domain.domain,
                                   reverse('Main:campaign_target_click',
                                           kwargs={'targetid': targetid})
                                   )
        else:
            linkurl = "%s%s" % (self.host_domain.domain,
                                reverse('Main:campaign_target_click',
                                        kwargs={'targetid': targetid})
                                )
        imgurl = "http://%s%s" % (self.host_domain.domain,
                                  reverse('Main:campaign_target_openmail',
                                          kwargs={'targetid': targetid})
                                  )

        html_content = base_mail["text"].replace("FIXMEURL", linkurl)

        if self.enable_mail_tracker:
            html_content = html_content.replace("FIXMEMAILTRACKER", imgurl)

        if target:
            for att in target.attributes.all():
                att.key = re.sub(r'ORDN-[0-9]{3}-','',att.key)
                html_content = html_content.replace("($%s$)" % (att.key), att.value)

        email.attach_alternative(html_content, "text/html")
        email.mixed_subtype = 'related'
        for img in base_mail["imgs"]:
            email.attach(img)

        email.extra_headers = {settings.PHISHING_MAIL_HEADER: "[%s]" % targetid}
        email.send(fail_silently=False)

    def __sendemailwithattachment(self,
                                  recipient,
                                  targetid,
                                  connec,
                                  mail_content,
                                  attachment_content):
        base_mail = mail_content
        attachment_content = attachment_content
        try:
            target = AnonymousTarget.objects.get(uniqueid=targetid)
        except ObjectDoesNotExist:
            target = None
        mail_content = "Read this mail with a HTML compatible client"
        sender = self.from_name + '@' + self.from_domain.domain
        if self.display_name != "":
            from_mail = "%s <%s>" % (self.display_name, sender)
        else:
            from_mail = sender

        email = mail.EmailMultiAlternatives(self.mail_template.title,
                                            mail_content,
                                            from_mail,
                                            [recipient],
                                            connection=connec)

        imgurl_mail = "http://%s%s" % (self.from_domain.domain,
                                       reverse('Main:campaign_target_openmail',
                                               kwargs={'targetid': targetid})
                                       )
        imgurl_attachment = "http://%s%s" % (self.from_domain.domain,
                                             reverse('Main:campaign_target_openattachment',
                                                     kwargs={'targetid': targetid})
                                             )

        html_content = base_mail["text"]
        mhtml_attach = attachment_content
        if self.enable_mail_tracker:
            html_content = html_content.replace("FIXMEMAILTRACKER", imgurl_mail)
        if self.enable_attachment_tracker:
            mhtml_attach = mhtml_attach.replace("FIXMEDOCTRACKER", imgurl_attachment)

        if target:
            for att in target.attributes.all():
                att.key = re.sub(r'ORDN-[0-9]{3}-','',att.key)
                html_content = html_content.replace("($%s$)" % (att.key), att.value)
        email.attach_alternative(html_content, "text/html")
        temp = "%s.doc" % (self.attachment_template.title)
        filename = Header(temp, 'utf-8').encode()
        email.attach(filename=filename,
                     content=mhtml_attach.encode("utf-8"),
                     mimetype="application/msword")
        email.mixed_subtype = 'related'
        for img in base_mail["imgs"]:
            email.attach(img)
        email.extra_headers = {settings.PHISHING_MAIL_HEADER: "[%s]" % targetid}
        email.send(fail_silently=False)

    def start(self):
        logger = logging.getLogger(__name__)
        if self.status != "2":
            self.status = "2"
        attachment_content = self.__buildattachment()
        self.targets_count = self.count_targets()
        self.save()
        try:
            send_alert_new_campaign(self.targets_count,
                                    self.author.user.email,
                                    self.mail_template.title,
                                    "%s@%s" % (self.from_name, self.from_domain.domain)
                                    )
        except SMTPRecipientsRefused:
            pass
        targetlists = self.targets.all()
        connec = mail.get_connection(fail_silently=False,timeout=15)
        connec.open()
        mail_content = self.__buildemail()
        for targetlist in targetlists:
            targets = targetlist.targets.all()
            for target in targets:
                newAnon = AnonymousTarget()
                newAnon.save()
                newAnon.importAttributes(target)
                try:
                    if self.campaign_type in ["1", "3", "4"]:
                        self.__sendemail(target.mail_address,
                                         newAnon.uniqueid,
                                         connec,
                                         mail_content
                                         )
                    elif self.campaign_type == "2":
                        self.__sendemailwithattachment(target.mail_address,
                                                       newAnon.uniqueid,
                                                       connec,
                                                       mail_content,
                                                       attachment_content
                                                       )
                except SMTPRecipientsRefused:
                    pass
                except SMTPServerDisconnected:
                    logger.error("Timeout connecting to SMTP server")
                    logger.info("Pause for a little while before opening a new connection")
                    sleep(60)
                    try:
                        connec = mail.get_connection(fail_silently=False,timeout=15)
                        connec.open()
                        if self.campaign_type in ["1", "3", "4"]:
                            self.__sendemail(target.mail_address,
                                             newAnon.uniqueid,
                                             connec,
                                             mail_content
                                             )
                        elif self.campaign_type == "2":
                            self.__sendemailwithattachment(target.mail_address,
                                                           newAnon.uniqueid,
                                                           connec,
                                                           mail_content,
                                                           attachment_content
                                                           )
                    except SMTPServerDisconnected:
                        logger.error("Failed to reopen a connection to SMTP server. Giving up")
                        break
                    else:
                        logger.info("New connection is working")
                        newAnon.mail_sent_time = datetime.datetime.now(tz=get_current_timezone())
                        self.anonymous_targets.add(newAnon)
                        newAnon.save()
                else:
                    newAnon.mail_sent_time = datetime.datetime.now(tz=get_current_timezone())
                    self.anonymous_targets.add(newAnon)
                    newAnon.save()
        connec.close()
        return True

    def stop(self):
        self.status = "3"
        self.save()
        return True

    def test(self, recipient):
        connec = mail.get_connection()
        connec.open()
        mail_content = self.__buildemail()
        attachment_content = self.__buildattachment()
        if self.campaign_type == "1" or self.campaign_type == "3" or self.campaign_type == "4":
            self.__sendemail(recipient, self.testid, connec, mail_content)
        elif self.campaign_type == "2":
            self.__sendemailwithattachment(recipient,
                                           self.testid,
                                           connec,
                                           mail_content,
                                           attachment_content
                                           )
        connec.close()
        return True

    def generate_results_xlsx(self):
        targets = self.anonymous_targets.all()
        wb = Workbook()
        ws = wb.active
        tags = set()
        dest_filename = 'results/' + str(self.id) + '.xlsx'

        for target in targets:
            attributes = target.attributes.all()
            for attribute in attributes:
                tags.add(attribute.key)

        header = []
        header.append("id")
        header.append("mail sent time")
        header.append("mail opened")
        header.append("mail opened time")
        if self.campaign_type == "1" or self.campaign_type == "3" or self.campaign_type == "4":
            header.append("link clicked")
            header.append("link clicked time")
            if self.campaign_type == "3":
                header.append("form submitted")
                header.append("form submitted time")
            if self.campaign_type == "4":
                header.append("img clicked")
                header.append("img clicked time")
        else:
            header.append("attachement opened")
            header.append("attachement opened time")
        header.append("reported")
        header.append("reported time")

        for tag in sorted(tags):
            header.append(re.sub(r'ORDN-[0-9]{3}-','',tag))

        ft = Font(bold=True)
        al = Alignment(horizontal="center", vertical="center")
        al2 = Alignment(horizontal="left", vertical="center")
        column = 1
        for cell in header:
            ws.cell(row=1, column=column, value=cell)
            ws.cell(row=1, column=column).font = ft
            ws.cell(row=1, column=column).alignment = al
            column += 1

        row = 2
        for target in targets:
            values = []

            if target.mail_sent_time:
                mail_sent_time = target.mail_sent_time.strftime('%Y-%m-%d %H:%M:%S')
            else:
                mail_sent_time = "N/A"

            if target.mail_opened:
                mail_opened = "yes"
            else:
                mail_opened = "no"

            if target.mail_opened_time:
                mail_opened_time = target.mail_opened_time.strftime('%Y-%m-%d %H:%M:%S')
            else:
                mail_opened_time = "N/A"

            if target.link_clicked:
                link_clicked = "yes"
            else:
                link_clicked = "no"

            if target.link_clicked_time:
                link_clicked_time = target.link_clicked_time.strftime('%Y-%m-%d %H:%M:%S')
            else:
                link_clicked_time = "N/A"

            if target.attachment_opened:
                attachment_opened = "yes"
            else:
                attachment_opened = "no"

            if target.attachment_opened_time:
                attachment_opened_time = target.attachment_opened_time.strftime('%Y-%m-%d %H:%M:%S')
            else:
                attachment_opened_time = "N/A"

            if target.form_submitted:
                form_submitted = "yes"
            else:
                form_submitted = "no"

            if target.form_submitted_time:
                form_submitted_time = target.form_submitted_time.strftime('%Y-%m-%d %H:%M:%S')
            else:
                form_submitted_time = "N/A"

            if target.reported:
                reported = "yes"
            else:
                reported = "no"

            if target.reported_time:
                reported_time = target.reported_time.strftime('%Y-%m-%d %H:%M:%S')
            else:
                reported_time = "N/A"

            for tag in sorted(tags):
                att = target.attributes.filter(key=tag)
                if att:
                    values.append(att[0].value)
                else:
                    values.append("N/A")

            column = 1
            ws.cell(row=row, column=column, value=target.uniqueid)
            ws.cell(row=row, column=column).font = ft
            ws.cell(row=row, column=column).alignment = al2
            column += 1
            ws.cell(row=row, column=column, value=mail_sent_time)
            ws.cell(row=row, column=column).alignment = al
            column += 1
            ws.cell(row=row, column=column, value=mail_opened)
            ws.cell(row=row, column=column).alignment = al
            column += 1
            ws.cell(row=row, column=column, value=mail_opened_time)
            ws.cell(row=row, column=column).alignment = al
            column += 1
            if self.campaign_type == "1" or self.campaign_type == "3" or self.campaign_type == "4":
                ws.cell(row=row, column=column, value=link_clicked)
                ws.cell(row=row, column=column).alignment = al
                column += 1
                ws.cell(row=row, column=column, value=link_clicked_time)
                ws.cell(row=row, column=column).alignment = al
                column += 1
                if self.campaign_type == "3" or self.campaign_type == "4":
                    ws.cell(row=row, column=column, value=form_submitted)
                    ws.cell(row=row, column=column).alignment = al
                    column += 1
                    ws.cell(row=row, column=column, value=form_submitted_time)
                    ws.cell(row=row, column=column).alignment = al
                    column += 1
            else:
                ws.cell(row=row, column=column, value=attachment_opened)
                ws.cell(row=row, column=column).alignment = al
                column += 1
                ws.cell(row=row, column=column, value=attachment_opened_time)
                ws.cell(row=row, column=column).alignment = al
                column += 1
            ws.cell(row=row, column=column, value=reported)
            ws.cell(row=row, column=column).alignment = al
            column += 1
            ws.cell(row=row, column=column, value=reported_time)
            ws.cell(row=row, column=column).alignment = al
            column += 1
            for val in values:
                ws.cell(row=row, column=column, value=val)
                ws.cell(row=row, column=column).alignment = al
                column += 1
            row += 1

        c = ws['B2']
        ws.freeze_panes = c
        wb.save(filename = dest_filename)
        return True

    def count_targets(self):
        lists = self.targets.all()
        total = 0
        for mylist in lists:
            total += mylist.targets.count()
        return total

    def links_clicked(self):
        count = self.anonymous_targets.filter(link_clicked=True).count()
        total = self.targets_count
        if total == 0:
            res = "0"
        else:
            res = "%s (%s%s)" % (count, int(float(count) / float(total) * 100), '%')
        return res

    def mails_reported(self):
        count = self.anonymous_targets.filter(reported=True).count()
        total = self.targets_count
        if total == 0:
            res = "0"
        else:
            res = "%s (%s%s)" % (count, int(float(count) / float(total) * 100), '%')
        return res

    def mails_open(self):
        count = self.anonymous_targets.filter(mail_opened=True).count()
        total = self.targets_count
        if total == 0:
            res = "0"
        else:
            res = "%s (%s%s)" % (count, int(float(count) / float(total) * 100), '%')
        return res

    def mails_sent(self):
        count = self.anonymous_targets.count()
        total = self.targets_count
        if total == 0:
            res = "0"
        else:
            res = "%s (%s%s)" % (count, int(float(count) / float(total) * 100), '%')
        return res

    def attachments_open(self):
        count = self.anonymous_targets.filter(attachment_opened=True).count()
        total = self.targets_count
        if total == 0:
            res = "0"
        else:
            res = "%s (%s%s)" % (count, int(float(count) / float(total) * 100), '%')
        return res

    def forms_submitted(self):
        count = self.anonymous_targets.filter(form_submitted=True).count()
        total = self.targets_count
        if total == 0:
            res = "0"
        else:
            res = "%s (%s%s)" % (count, int(float(count) / float(total) * 100), '%')
        return res
